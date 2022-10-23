import openpyxl 

class parseToExcel:
    def __init__(self):
        self.EXCEL_NAME = "database.xlsx"
        self.TXT_NAME = "link.txt"

        self.EXCEL = None
        self.TXT_FILE = None

        self.lineCount = 0

        self.sheet = None

        self.data = {
            "year": None,
            "term": None,
            "course": None,
            "assignment": None,
            "weight": None,
            "grade": None
        }

        self.accessFiles()
        self.getData()
        if self.data['year'] and self.data['term']:
            self.accessSheet()
        elif self.data['year']:
            print("Year, but no term")
        elif self.data['term']:
            print("Term, but no year")
        else:
            print(self.data['year'],self.data['term'])
        if (self.data['course']):
            self.plugData()
        self.saveExcel()

    def accessFiles(self):
        print("Getting files...")
        opened = False
        # Abrir/crear archivo .txt
        while not opened:
            try:
                self.TXT_FILE = open(self.TXT_NAME, 'r')
                opened = True
                print("Got the .txt!")
            except:
                self.TXT_FILE = open(self.TXT_NAME, 'w')
                self.TXT_FILE.close()


        # Abrir/crear archivo .xlsx
        try:
            self.EXCEL = openpyxl.load_workbook(filename = self.EXCEL_NAME)
            print("Got the workbook!")
        except:
            self.EXCEL = openpyxl.Workbook()
            self.saveExcel()

    def accessSheet(self):
        print("Accessing sheet...")        
        sheetname = self.data['term'] + self.data['year']
        try:
            self.sheet = self.EXCEL[sheetname]
            print("Found it!")
        except:
            # print(type(self.data["year"]))
            print("Created new sheet")
            self.sheet = self.EXCEL.create_sheet(sheetname)

        self.saveExcel()

    def getData(self):
        print("Getting Data...")
        counter = 1
        rawData = []
        for line in self.TXT_FILE:
            rawData.append(line.strip())
            print("Got line", counter)
            counter += 1
        
        self.lineCount = len(rawData)
        
        dataKeys = list(self.data.keys())

        for i in range(len(rawData)):
            self.data[dataKeys[i]] = rawData[i]

    def plugData(self):
        print("Plugging Data...")
        assignmentRow = 2
        gap = 3
        # Get to first blank column (row 1)
        for col in self.EXCEL.iter_cols():
            if (col[1].value == self.data['course']):
                currentColumn = col
                break
            elif (not col[1].value):
                self.EXCEL.insert_cols(col, amount = 1)
                currentColumn = col
                # Write course name
                currentColumn[1].value = self.data['course']

        # Start of pattern (row n)
        # Assignment name -> (row n)
        if (self.data['assignment']):
            print("There's data in assignment")
            
            # Agregar fila si no existe
            # Guardar fila si ya existe
            found = currentColumn[assignmentRow].value == self.data['assignment']
            while not found:
                assignmentRow += gap
                found = currentColumn[assignmentRow].value == self.data['assignment']
                if not currentColumn[assignmentRow].value:
                    self.EXCEL.insert_rows(self.EXCEL[assignmentRow], amount = gap)
                    currentColumn[assignmentRow].value = self.data['assignment']
                    found = True
                
        
            # Assignment weight -> (row n+1)
            currentColumn[assignmentRow+1].value = self.data['weight']

            # Assignment grade -> (row n+2)
            currentColumn[assignmentRow+2].value = self.data['grade']
        
        return
        
        

    def saveExcel(self):
        self.EXCEL.save(filename = self.EXCEL_NAME)


# Fila 1 de .txt -> año (Hoja)
# Fila 2 de .txt -> Term (1ra fila de xlsx)
# Fila 3 de .txt -> Course (2da fila de xlsx)
# Fila 4 de .txt -> Assignment (3ra fila de xlsx)
# Fila 5 de .txt -> Weight (4ta fila de xlsx)
# Fila 6 de .txt -> Grade (5ta fila de xlsx)

parser = parseToExcel()